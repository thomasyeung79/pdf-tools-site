from __future__ import annotations

import shutil
import subprocess
import uuid
from pathlib import Path
from threading import Lock
from urllib.parse import unquote, urlparse

import pandas as pd
import requests
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlsxImage
from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
STORAGE_DIR = BASE_DIR / "storage"
UPLOAD_DIR = STORAGE_DIR / "uploads"
OUTPUT_DIR = STORAGE_DIR / "outputs"
MAX_FILE_SIZE = 200 * 1024 * 1024  # 200 MB

for directory in (UPLOAD_DIR, OUTPUT_DIR):
    directory.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="免费 PDF 工具站 · 加强版")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# ── 后台任务状态 ──────────────────────────────────────────────
_task_status: dict[str, dict] = {}
_task_lock = Lock()


def _set_task(task_id: str, status: str, detail: str = "", progress: int = 0):
    with _task_lock:
        _task_status[task_id] = {"status": status, "detail": detail, "progress": progress}


def _get_task(task_id: str) -> dict | None:
    with _task_lock:
        return _task_status.get(task_id)


# ── 首页 ──────────────────────────────────────────────────────


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return (STATIC_DIR / "index.html").read_text(encoding="utf-8")


@app.get("/api/config")
def config():
    return {
        "output_dir": str(OUTPUT_DIR.resolve()),
        "max_file_size": MAX_FILE_SIZE,
        "max_file_size_display": f"{MAX_FILE_SIZE // (1024*1024)} MB",
    }


@app.get("/api/task/{task_id}")
def task_status(task_id: str):
    status = _get_task(task_id)
    if not status:
        raise HTTPException(status_code=404, detail="任务不存在")
    if status["status"] == "error":
        raise HTTPException(status_code=400, detail=status["detail"])
    return status


# ── 工具函数 ──────────────────────────────────────────────────


async def _save_upload(file: UploadFile, *extensions: str) -> Path:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in extensions:
        raise HTTPException(status_code=400, detail=f"文件类型不支持，请上传：{'、'.join(extensions)}")

    path = UPLOAD_DIR / f"{uuid.uuid4().hex}_{_safe_filename(file.filename or 'upload')}"
    size = 0
    with path.open("wb") as output:
        while chunk := await file.read(1024 * 1024):
            size += len(chunk)
            if size > MAX_FILE_SIZE:
                path.unlink(missing_ok=True)
                raise HTTPException(status_code=413, detail=f"文件太大，最大允许 {MAX_FILE_SIZE // (1024*1024)} MB。")
            output.write(chunk)
    return path


def _download_response(path: Path, filename: str, media_type: str) -> FileResponse:
    return FileResponse(
        path,
        filename=filename,
        media_type=media_type,
        headers={"X-Output-Path": str(path.resolve())},
    )


def _safe_filename(name: str) -> str:
    blocked = '<>:"/\\|?*'
    cleaned = "".join("_" if char in blocked else char for char in name).strip()
    return cleaned or "download"


# ── PDF 链接下载 ──────────────────────────────────────────────


@app.post("/api/download-pdf")
def download_pdf(url: str = Form(...), filename: str = Form("download.pdf")):
    pdf_url = _extract_pdf_url(url)
    safe_name = _safe_filename(filename or Path(urlparse(pdf_url).path).name or "download.pdf")
    if not safe_name.lower().endswith(".pdf"):
        safe_name += ".pdf"

    output_path = OUTPUT_DIR / f"{uuid.uuid4().hex}_{safe_name}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125 Safari/537.36",
        "Accept": "application/pdf,application/octet-stream,*/*",
        "Referer": f"{urlparse(url).scheme}://{urlparse(url).netloc}/" if urlparse(url).netloc else "",
    }

    try:
        with requests.get(pdf_url, headers=headers, stream=True, timeout=60, allow_redirects=True) as response:
            response.raise_for_status()
            content_length = response.headers.get("content-length")
            if content_length and int(content_length) > MAX_FILE_SIZE:
                raise HTTPException(status_code=413, detail=f"远程文件太大（{int(content_length) // (1024*1024)} MB），最大允许 {MAX_FILE_SIZE // (1024*1024)} MB。")
            with output_path.open("wb") as file:
                downloaded = 0
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        downloaded += len(chunk)
                        if downloaded > MAX_FILE_SIZE:
                            output_path.unlink(missing_ok=True)
                            raise HTTPException(status_code=413, detail=f"文件太大，最大允许 {MAX_FILE_SIZE // (1024*1024)} MB。")
                        file.write(chunk)
    except requests.RequestException as exc:
        raise HTTPException(status_code=400, detail=f"下载失败：{exc}") from exc

    if output_path.read_bytes()[:4] != b"%PDF":
        output_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="下载到的内容不是 PDF，可能需要登录、验证码或权限。")

    return _download_response(output_path, safe_name, "application/pdf")


def _extract_pdf_url(url: str) -> str:
    parsed = urlparse(url)
    if "course_pdf=" not in url:
        return url

    query_source = parsed.fragment if parsed.fragment else parsed.query
    marker = "course_pdf="
    start = query_source.find(marker)
    if start == -1:
        return url
    value = query_source[start + len(marker) :].split("&", 1)[0]
    return unquote(value)


# ── PDF 转 Word ──────────────────────────────────────────────


@app.post("/api/pdf-to-word")
async def pdf_to_word(file: UploadFile = File(...)):
    input_path = await _save_upload(file, ".pdf")
    output_path = OUTPUT_DIR / f"{input_path.stem}.docx"

    try:
        from pdf2docx import Converter

        converter = Converter(str(input_path))
        converter.convert(str(output_path), start=0, end=None)
        converter.close()
    except ImportError:
        try:
            _convert_pdf_pages_to_docx(input_path, output_path)
        except Exception as exc:
            raise HTTPException(status_code=501, detail=f"缺少 pdf2docx，且图片版 Word 备用转换失败：{exc}") from exc
    except Exception as exc:
        try:
            _convert_pdf_pages_to_docx(input_path, output_path)
        except Exception:
            raise HTTPException(status_code=400, detail=f"PDF 转 Word 失败：{exc}") from exc

    return _download_response(output_path, output_path.name, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


def _convert_pdf_pages_to_docx(input_path: Path, output_path: Path) -> None:
    try:
        import fitz
        from docx import Document
        from docx.shared import Inches
    except ImportError as exc:
        raise RuntimeError("缺少 pymupdf 或 python-docx，请运行 pip install pymupdf python-docx。") from exc

    image_dir = OUTPUT_DIR / f"{input_path.stem}_{uuid.uuid4().hex}_pages"
    image_dir.mkdir(parents=True, exist_ok=True)

    document = Document()
    section = document.sections[0]
    section.top_margin = Inches(0.35)
    section.bottom_margin = Inches(0.35)
    section.left_margin = Inches(0.35)
    section.right_margin = Inches(0.35)
    max_width = section.page_width - section.left_margin - section.right_margin

    pdf = fitz.open(str(input_path))
    try:
        for page_number, page in enumerate(pdf, start=1):
            image_path = image_dir / f"page_{page_number}.png"
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            pixmap.save(str(image_path))
            if page_number > 1:
                document.add_page_break()
            document.add_picture(str(image_path), width=max_width)
    finally:
        pdf.close()

    document.save(str(output_path))


# ── Word 转 PDF ──────────────────────────────────────────────


@app.post("/api/word-to-pdf")
async def word_to_pdf(file: UploadFile = File(...)):
    input_path = await _save_upload(file, ".docx", ".doc")
    result = (
        subprocess.run(
            [_find_soffice(), "--headless", "--convert-to", "pdf", "--outdir", str(OUTPUT_DIR), str(input_path)],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if _find_soffice()
        else None
    )

    output_path = OUTPUT_DIR / f"{input_path.stem}.pdf"
    if result and result.returncode == 0 and output_path.exists():
        return _download_response(output_path, output_path.name, "application/pdf")

    try:
        _convert_word_with_ms_word(input_path, output_path)
    except Exception as exc:
        libreoffice_error = result.stderr or result.stdout if result else ""
        message = "Word 转 PDF 需要安装 LibreOffice，或在 Windows 上安装 Microsoft Word。"
        if libreoffice_error:
            message += f" LibreOffice 错误：{libreoffice_error}"
        raise HTTPException(status_code=501, detail=f"{message} Word 自动导出错误：{exc}") from exc

    return _download_response(output_path, output_path.name, "application/pdf")


def _find_soffice() -> str | None:
    command = shutil.which("soffice") or shutil.which("libreoffice")
    if command:
        return command

    common_paths = [
        Path("C:/Program Files/LibreOffice/program/soffice.exe"),
        Path("C:/Program Files (x86)/LibreOffice/program/soffice.exe"),
    ]
    for path in common_paths:
        if path.exists():
            return str(path)
    return None


def _convert_word_with_ms_word(input_path: Path, output_path: Path) -> None:
    try:
        import pythoncom
        from win32com.client import DispatchEx
    except ImportError as exc:
        raise RuntimeError("缺少 pywin32，请运行 pip install pywin32。") from exc

    pythoncom.CoInitialize()
    word = None
    document = None
    try:
        word = DispatchEx("Word.Application")
        word.Visible = False
        document = word.Documents.Open(str(input_path.resolve()))
        document.SaveAs(str(output_path.resolve()), FileFormat=17)
    finally:
        if document is not None:
            document.Close(False)
        if word is not None:
            word.Quit()
        pythoncom.CoUninitialize()


# ── Excel 转 PDF ─────────────────────────────────────────────


@app.post("/api/excel-to-pdf")
async def excel_to_pdf(file: UploadFile = File(...)):
    input_path = await _save_upload(file, ".xlsx", ".xls", ".csv")
    output_path = OUTPUT_DIR / f"{input_path.stem}.pdf"
    rows = _read_table_rows(input_path)

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    table = Table(rows[:200], repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ])
    )
    document.build([table])

    return _download_response(output_path, output_path.name, "application/pdf")


# ── Excel 转图片 ─────────────────────────────────────────────


@app.post("/api/excel-to-image")
async def excel_to_image(file: UploadFile = File(...)):
    input_path = await _save_upload(file, ".xlsx", ".xls", ".csv")
    output_path = OUTPUT_DIR / f"{input_path.stem}.png"
    rows = _read_table_rows(input_path)[:80]

    cell_w = 160
    cell_h = 34
    width = min(max(len(rows[0]) * cell_w, 640), 2400)
    height = min(max(len(rows) * cell_h, 240), 3200)
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()

    for row_index, row in enumerate(rows):
        y = row_index * cell_h
        fill = "#0f766e" if row_index == 0 else ("#f8fafc" if row_index % 2 else "#ffffff")
        text_fill = "white" if row_index == 0 else "#0f172a"
        for col_index, value in enumerate(row):
            x = col_index * cell_w
            if x >= width:
                break
            draw.rectangle((x, y, x + cell_w, y + cell_h), fill=fill, outline="#cbd5e1")
            draw.text((x + 8, y + 10), str(value)[:22], fill=text_fill, font=font)

    image.save(output_path)
    return _download_response(output_path, output_path.name, "image/png")


# ── 图片转 Excel ─────────────────────────────────────────────


@app.post("/api/image-to-excel")
async def image_to_excel(file: UploadFile = File(...)):
    input_path = await _save_upload(file, ".png", ".jpg", ".jpeg", ".webp")
    output_path = OUTPUT_DIR / f"{input_path.stem}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Image"
    sheet["A1"] = "图片已嵌入。若要识别表格文字，需要在服务器安装 OCR，例如 Tesseract。"
    img = XlsxImage(str(input_path))
    img.width = min(img.width, 900)
    img.height = min(img.height, 1200)
    sheet.add_image(img, "A3")
    workbook.save(output_path)

    return _download_response(
        output_path, output_path.name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ── 图片提取文字 ─────────────────────────────────────────────


@app.post("/api/image-to-text")
async def image_to_text(file: UploadFile = File(...), language: str = Form("ch")):
    input_path = await _save_upload(file, ".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff")
    output_path = OUTPUT_DIR / f"{input_path.stem}_text.txt"

    try:
        text = await run_in_threadpool(_extract_text_from_image, input_path, language)
    except Exception as exc:
        raise HTTPException(status_code=501, detail=str(exc)) from exc

    output_path.write_text(text.strip() + "\n", encoding="utf-8")
    return _download_response(output_path, output_path.name, "text/plain; charset=utf-8")


def _extract_text_from_image(input_path: Path, language: str) -> str:
    try:
        from rapidocr_onnxruntime import RapidOCR

        engine = RapidOCR()
        result, _ = engine(str(input_path))
        lines = [item[1] for item in result or [] if len(item) >= 2 and item[1]]
        if lines:
            return "\n".join(lines)
    except ImportError:
        pass

    try:
        import pytesseract
    except ImportError as exc:
        raise RuntimeError(
            "图片提取文字需要 OCR 依赖。请先运行 pip install rapidocr-onnxruntime，或安装 Tesseract 后运行 pip install pytesseract。"
        ) from exc

    lang_map = {"ch": "chi_sim+eng", "en": "eng"}
    try:
        return pytesseract.image_to_string(Image.open(input_path), lang=lang_map.get(language, "chi_sim+eng"))
    except Exception as exc:
        raise RuntimeError(f"OCR 识别失败：{exc}") from exc


# ── 合并 PDF ────────────────────────────────────────────────


@app.post("/api/merge-pdf")
async def merge_pdf(files: list[UploadFile] = File(...)):
    if len(files) < 2:
        raise HTTPException(status_code=400, detail="请至少上传 2 个 PDF 文件。")

    writer = PdfWriter()
    saved_files = [await _save_upload(file, ".pdf") for file in files]
    output_path = OUTPUT_DIR / f"merged_{uuid.uuid4().hex}.pdf"

    try:
        for path in saved_files:
            reader = PdfReader(str(path))
            for page in reader.pages:
                writer.add_page(page)
        with output_path.open("wb") as output:
            writer.write(output)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"合并 PDF 失败：{exc}") from exc

    return _download_response(output_path, "merged.pdf", "application/pdf")


# ── 拆分 PDF ────────────────────────────────────────────────


@app.post("/api/split-pdf")
async def split_pdf(file: UploadFile = File(...), pages: str = Form("")):
    input_path = await _save_upload(file, ".pdf")
    output_path = OUTPUT_DIR / f"{input_path.stem}_split.pdf"

    try:
        reader = PdfReader(str(input_path))
        selected_pages = _parse_page_ranges(pages, len(reader.pages))
        writer = PdfWriter()
        for page_index in selected_pages:
            writer.add_page(reader.pages[page_index])
        with output_path.open("wb") as output:
            writer.write(output)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"拆分 PDF 失败：{exc}") from exc

    return _download_response(output_path, f"{Path(file.filename or 'split').stem}_split.pdf", "application/pdf")


def _parse_page_ranges(value: str, page_count: int) -> list[int]:
    if page_count <= 0:
        raise ValueError("PDF 没有可拆分的页面。")

    text = value.strip()
    if not text:
        return list(range(page_count))

    selected: list[int] = []
    for part in text.replace("，", ",").split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start_text, end_text = part.split("-", 1)
            start = int(start_text)
            end = int(end_text)
            if start > end:
                raise ValueError("页码范围格式错误，例如 1-3。")
            selected.extend(range(start - 1, end))
        else:
            selected.append(int(part) - 1)

    if not selected:
        raise ValueError("请输入要拆分的页码，例如 1,3-5。")
    if min(selected) < 0 or max(selected) >= page_count:
        raise ValueError(f"页码超出范围，这个 PDF 共 {page_count} 页。")

    return list(dict.fromkeys(selected))


# ═══════════════════════════ 新增功能 ═══════════════════════════


# ── 压缩 PDF ────────────────────────────────────────────────


@app.post("/api/compress-pdf")
async def compress_pdf(file: UploadFile = File(...), quality: str = Form("medium")):
    """通过降低图片质量来压缩 PDF。"""
    input_path = await _save_upload(file, ".pdf")
    output_path = OUTPUT_DIR / f"{input_path.stem}_compressed.pdf"
    quality_map = {"high": 150, "medium": 100, "low": 60}

    try:
        import fitz
    except ImportError as exc:
        raise HTTPException(status_code=501, detail="压缩 PDF 需要 PyMuPDF，请运行 pip install pymupdf。") from exc

    dpi = quality_map.get(quality, 100)
    pdf = fitz.open(str(input_path))
    new_pdf = fitz.open()
    page_count = pdf.page_count

    try:
        for idx in range(page_count):
            page = pdf[idx]
            # 渲染页面为图片再插回，丢弃原始向量信息以减小体积
            zoom = dpi / 72
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
            img_bytes = pix.tobytes("jpeg")
            rect = page.rect
            new_page = new_pdf.new_page(width=rect.width, height=rect.height)
            new_page.insert_image(rect, stream=img_bytes)
    finally:
        pdf.close()

    new_pdf.save(str(output_path), deflate=True, garbage=3, clean=True)
    new_pdf.close()

    original_size = input_path.stat().st_size
    compressed_size = output_path.stat().st_size
    ratio = (1 - compressed_size / original_size) * 100 if original_size else 0

    # 如果压缩后反而更大，返回原版的复制
    if compressed_size >= original_size:
        output_path.unlink(missing_ok=True)
        shutil.copy2(str(input_path), str(output_path))
        ratio = 0

    return _download_response(output_path, output_path.name, "application/pdf")


# ── PDF 转图片 ──────────────────────────────────────────────


@app.post("/api/pdf-to-images")
async def pdf_to_images(file: UploadFile = File(...), dpi: int = Form(150)):
    """将 PDF 每页转为 PNG 图片，打包为 ZIP 下载。"""
    input_path = await _save_upload(file, ".pdf")

    try:
        import fitz
        import zipfile
    except ImportError as exc:
        raise HTTPException(status_code=501, detail="PDF 转图片需要 PyMuPDF，请运行 pip install pymupdf。") from exc

    image_dir = OUTPUT_DIR / f"{input_path.stem}_images_{uuid.uuid4().hex}"
    image_dir.mkdir(parents=True, exist_ok=True)

    pdf = fitz.open(str(input_path))
    try:
        for idx in range(pdf.page_count):
            page = pdf[idx]
            zoom = dpi / 72
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
            img_path = image_dir / f"page_{idx + 1:04d}.png"
            pix.save(str(img_path))
    finally:
        pdf.close()

    zip_path = OUTPUT_DIR / f"{input_path.stem}_images_{uuid.uuid4().hex}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for img_file in sorted(image_dir.iterdir()):
            zf.write(img_file, img_file.name)

    # 清理临时图片目录
    shutil.rmtree(image_dir, ignore_errors=True)

    return _download_response(zip_path, zip_path.name, "application/zip")


# ── 图片合成 PDF ────────────────────────────────────────────


@app.post("/api/images-to-pdf")
async def images_to_pdf(files: list[UploadFile] = File(...)):
    """将多张图片合并为一个 PDF。"""
    if len(files) < 1:
        raise HTTPException(status_code=400, detail="请至少上传 1 张图片。")

    saved_files = []
    for f in files:
        p = await _save_upload(f, ".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff")
        saved_files.append(p)

    output_path = OUTPUT_DIR / f"images_combined_{uuid.uuid4().hex}.pdf"

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        a4_w, a4_h = A4
        c = canvas.Canvas(str(output_path), pagesize=A4)
        for img_path in saved_files:
            img = Image.open(img_path)
            # 保持比例缩放
            iw, ih = img.size
            scale = min(a4_w / iw, a4_h / ih) * 0.92
            dw, dh = iw * scale, ih * scale
            x = (a4_w - dw) / 2
            y = (a4_h - dh) / 2
            c.drawImage(str(img_path), x, y, width=dw, height=dh)
            c.showPage()
        c.save()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"合成 PDF 失败：{exc}") from exc

    return _download_response(output_path, "images_combined.pdf", "application/pdf")


# ── PDF 信息查看 ────────────────────────────────────────────


@app.post("/api/pdf-info")
async def pdf_info(file: UploadFile = File(...)):
    """读取 PDF 元信息。"""
    input_path = await _save_upload(file, ".pdf")

    try:
        reader = PdfReader(str(input_path))
        info = reader.metadata
        meta = {}
        if info:
            for k, v in info:
                meta[k] = str(v) if not isinstance(v, str) else v

        # 每页信息
        pages_info = []
        for idx, page in enumerate(reader.pages, start=1):
            p = {"number": idx, "width": round(page.mediabox.width, 1), "height": round(page.mediabox.height, 1)}
            pages_info.append(p)

        data = {
            "filename": input_path.name,
            "page_count": len(reader.pages),
            "file_size_bytes": input_path.stat().st_size,
            "file_size_display": _format_size(input_path.stat().st_size),
            "pdf_version": reader.pdf_header,
            "metadata": meta,
            "pages": pages_info,
        }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 PDF 信息失败：{exc}") from exc

    return JSONResponse(content=data)


def _format_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024*1024):.1f} MB"
    return f"{size_bytes / (1024*1024*1024):.2f} GB"


# ── PDF 添加水印 ────────────────────────────────────────────


@app.post("/api/pdf-watermark")
async def pdf_watermark(
    file: UploadFile = File(...),
    text: str = Form("机密"),
    opacity: float = Form(0.3),
    angle: int = Form(45),
):
    """为 PDF 每页添加文字水印（对角铺满）。"""
    input_path = await _save_upload(file, ".pdf")
    output_path = OUTPUT_DIR / f"{input_path.stem}_watermarked.pdf"

    try:
        import fitz
    except ImportError as exc:
        raise HTTPException(status_code=501, detail="水印功能需要 PyMuPDF，请运行 pip install pymupdf。") from exc

    opacity = max(0.05, min(1.0, opacity))
    pdf = fitz.open(str(input_path))
    try:
        for page in pdf:
            rect = page.rect
            # 在页面上创建水印
            width, height = rect.width, rect.height
            shape = page.new_shape()
            shape.insert_text(
                fitz.Point(width * 0.1, height * 0.1),
                text,
                fontsize=max(width, height) / 12,
                rotate=angle,
                color=(0.5, 0.5, 0.5),
                opacity=opacity,
            )
            # 对角铺满多个水印
            for row in range(4):
                for col in range(4):
                    x = width * 0.05 + col * width * 0.25
                    y = height * 0.05 + row * height * 0.25
                    shape.insert_text(
                        fitz.Point(x, y),
                        text,
                        fontsize=max(width, height) / 14,
                        rotate=angle,
                        color=(0.5, 0.5, 0.5),
                        opacity=opacity,
                    )
            shape.commit()
    finally:
        pdf.close()

    # 重新打开保存
    pdf = fitz.open(str(input_path))
    pdf.save(str(output_path), deflate=True, garbage=3)
    pdf.close()

    return _download_response(output_path, output_path.name, "application/pdf")


# ── PDF 去水印 ──────────────────────────────────────────────


@app.post("/api/remove-watermark")
async def remove_watermark(
    file: UploadFile = File(...),
    watermark_text: str = Form(""),
    mode: str = Form("auto"),
):
    """移除 PDF 中的文字水印。

    - mode=text: 删除指定的文字（watermark_text 必填）
    - mode=auto: 自动检测跨页重复的文字水印
    - mode=image: 图片渲染模式（将 PDF 转为图片后擦除，需要 opencv-python）
    """
    input_path = await _save_upload(file, ".pdf")
    output_path = OUTPUT_DIR / f"{input_path.stem}_clean.pdf"

    try:
        import fitz
    except ImportError as exc:
        raise HTTPException(status_code=501, detail="去水印需要 PyMuPDF，请运行 pip install pymupdf。") from exc

    if mode == "image":
        # ── 图片渲染模式 ─────────────────────────────────────
        _remove_watermark_via_image(input_path, output_path)
    else:
        # ── PDF 内容流模式 ───────────────────────────────────
        pdf = fitz.open(str(input_path))
        try:
            if mode == "text" and watermark_text:
                _remove_specific_text_watermark(pdf, watermark_text)
            else:
                _auto_remove_watermark(pdf)
        finally:
            pdf.save(str(output_path), deflate=True, garbage=3, clean=True)
            pdf.close()

    original_size = input_path.stat().st_size
    new_size = output_path.stat().st_size

    # 如果处理后的文件比原文件大很多（缩水超过 50% 说明出了问题），返回警告
    if new_size > original_size * 1.5:
        # 可能有问题，但仍然返回文件
        pass

    return _download_response(output_path, output_path.name, "application/pdf")


def _remove_specific_text_watermark(pdf, watermark_text: str) -> int:
    """删除 PDF 中所有匹配指定文字的文本框。"""
    removed = 0
    for page in pdf:
        # 搜索所有匹配的文本位置
        instances = page.search_for(watermark_text)
        for inst in instances:
            page.add_redact_annot(inst, fill=(1, 1, 1))  # 白色填充覆盖
            removed += 1
        if instances:
            page.apply_redactions()
    return removed


def _auto_remove_watermark(pdf) -> int:
    """自动检测并移除跨页重复的文字水印。

    策略：
    1. 收集每页的文字位置和内容
    2. 找出在 >50% 页面相同位置出现的文字
    3. 对疑似水印的文字添加红色遮盖
    """
    import fitz
    from collections import defaultdict

    page_count = pdf.page_count
    if page_count == 0:
        return 0

    # ── 第 1 步：收集每页的文字区块 ────────────────────────
    pages_text_blocks = []
    for page in pdf:
        blocks = page.get_text("dict")["blocks"]
        text_spans = []
        for block in blocks:
            if block["type"] != 0:  # 不是文字块
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    text = span.get("text", "").strip()
                    if len(text) < 2:  # 忽略单字符
                        continue
                    bbox = span["bbox"]
                    size = round(span.get("size", 0), 1)
                    color = span.get("color", 0)
                    flags = span.get("flags", 0)
                    font = span.get("font", "")
                    # 记录位置（取中心点，四舍五入到 10px 粒度）
                    cx = round((bbox[0] + bbox[2]) / 2, -1)
                    cy = round((bbox[1] + bbox[3]) / 2, -1)
                    text_spans.append({
                        "text": text,
                        "bbox": bbox,
                        "cx": cx,
                        "cy": cy,
                        "size": size,
                        "color": color,
                        "font": font,
                        "flags": flags,
                    })
        pages_text_blocks.append(text_spans)

    if not any(pages_text_blocks):
        return 0

    # ── 第 2 步：按 (文字, 位置) 统计出现次数 ─────────────
    position_count = defaultdict(lambda: {"count": 0, "bbox": None, "color": None})
    for spans in pages_text_blocks:
        seen_this_page = set()
        for s in spans:
            key = (s["text"], s["cx"], s["cy"])
            if key not in seen_this_page:
                seen_this_page.add(key)
                position_count[key]["count"] += 1
                position_count[key]["bbox"] = s["bbox"]
                position_count[key]["color"] = s["color"]

    # ── 第 3 步：找出疑似水印的区域 ─────────────────────────
    threshold = max(2, page_count * 0.5)  # 超过半数的页面出现
    watermark_bboxes = []

    for key, info in position_count.items():
        if info["count"] < threshold:
            continue
        # 水印常见特征：灰色、较大字号、居中或重复排列
        bbox = info["bbox"]
        color = info["color"]
        # 灰色 (R=G=B) 的文本很可能是水印
        is_gray = isinstance(color, (int, float)) and color != 0
        is_light_gray = isinstance(color, (int, float)) and 0.3 <= color <= 0.8
        is_centered = False
        if bbox:
            w = bbox[2] - bbox[0]
            h = bbox[3] - bbox[1]
            # 很宽的文字块（居中的页眉/页脚水印）
            is_centered = w > 150

        if is_gray or is_light_gray or is_centered:
            watermark_bboxes.append(bbox)
        else:
            # 即使不满足特征，也加进来（宁可误杀）
            watermark_bboxes.append(bbox)

    # ── 第 4 步：不再对文字特征做二次判断，直接收集所有跨页重复区域 ──
    # 水印的核心特征就是"跨页同位置重复"，这已经足够
    watermark_bboxes = []
    for key, info in position_count.items():
        if info["count"] >= threshold and info["bbox"]:
            watermark_bboxes.append(info["bbox"])

    # ── 第 5 步：执行遮盖 ────────────────────────────────────
    removed = 0
    if watermark_bboxes:
        for page in pdf:
            for bbox in watermark_bboxes:
                page.add_redact_annot(bbox, fill=(1, 1, 1))
                removed += 1
            page.apply_redactions()

    return removed


def _remove_watermark_via_image(input_path: Path, output_path: Path):
    """图片渲染模式：将 PDF 每页转为图片，用 OpenCV 做水印擦除，再合成为 PDF。

    适用于水印已嵌入图片或无法通过文本方式移除的场景。
    需要安装 opencv-python 和 pymupdf。
    """
    import fitz

    try:
        import cv2
        import numpy as np
    except ImportError as exc:
        raise RuntimeError("图片模式去水印需要 opencv-python，请运行 pip install opencv-python numpy。") from exc

    pdf = fitz.open(str(input_path))
    new_pdf = fitz.open()

    try:
        for idx in range(pdf.page_count):
            page = pdf[idx]
            # 高分辨率渲染
            zoom = 2.0  # 144 DPI
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
            img_bytes = pix.tobytes("png")

            # 转为 OpenCV 格式
            nparr = np.frombuffer(img_bytes, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            if img is None:
                continue

            # 水印擦除策略：
            # 1. 转为灰度，用阈值检测低对比度的灰色文字区域
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            # 水印通常是浅灰色文字 → 用阈值提取
            _, mask = cv2.threshold(gray, 160, 200, cv2.THRESH_BINARY)
            mask = cv2.bitwise_not(mask)  # 反转，水印区域为白色

            # 扩张蒙版覆盖水印笔画
            kernel = np.ones((3, 3), np.uint8)
            mask = cv2.dilate(mask, kernel, iterations=1)

            # 用 Telea 算法修复
            result = cv2.inpaint(img, mask, 3, cv2.INPAINT_TELEA)

            # 写回新 PDF
            _, buffer = cv2.imencode(".png", result)
            rect = page.rect
            new_page = new_pdf.new_page(width=rect.width, height=rect.height)
            new_page.insert_image(rect, stream=buffer.tobytes())
    finally:
        pdf.close()

    new_pdf.save(str(output_path), deflate=True, garbage=3, clean=True)
    new_pdf.close()


@app.post("/api/rotate-pdf")
async def rotate_pdf(file: UploadFile = File(...), angle: int = Form(90)):
    """旋转 PDF 所有页面。"""
    input_path = await _save_upload(file, ".pdf")
    output_path = OUTPUT_DIR / f"{input_path.stem}_rotated.pdf"

    allowed = {90, 180, 270}
    if angle not in allowed:
        raise HTTPException(status_code=400, detail=f"旋转角度仅支持：{', '.join(map(str, allowed))} 度。")

    try:
        reader = PdfReader(str(input_path))
        writer = PdfWriter()
        for page in reader.pages:
            page.rotate(angle)
            writer.add_page(page)
        with output_path.open("wb") as f:
            writer.write(f)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"旋转 PDF 失败：{exc}") from exc

    return _download_response(output_path, output_path.name, "application/pdf")


# ── Excel / CSV 互转 ────────────────────────────────────────


@app.post("/api/excel-to-csv")
async def excel_to_csv(file: UploadFile = File(...)):
    """Excel 转 CSV (UTF-8)。"""
    input_path = await _save_upload(file, ".xlsx", ".xls", ".csv")
    output_path = OUTPUT_DIR / f"{input_path.stem}.csv"

    if input_path.suffix.lower() == ".csv":
        # 已经是 CSV，直接复制
        shutil.copy2(str(input_path), str(output_path))
    else:
        df = _read_dataframe(input_path)
        df.to_csv(output_path, index=False, encoding="utf-8-sig")

    return _download_response(output_path, output_path.name, "text/csv; charset=utf-8")


@app.post("/api/csv-to-excel")
async def csv_to_excel(file: UploadFile = File(...)):
    """CSV 转 Excel。"""
    input_path = await _save_upload(file, ".csv")
    output_path = OUTPUT_DIR / f"{input_path.stem}.xlsx"
    df = _read_dataframe(input_path)
    df.to_excel(output_path, index=False, engine="openpyxl")

    return _download_response(output_path, output_path.name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _read_dataframe(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, dtype=str).fillna("")
    else:
        return pd.read_excel(path, dtype=str, engine="openpyxl").fillna("")


# ── 通用表格读取 ────────────────────────────────────────────


def _read_table_rows(path: Path) -> list[list[str]]:
    if path.suffix.lower() == ".csv":
        dataframe = pd.read_csv(path, dtype=str).fillna("")
    else:
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = [[("" if cell is None else str(cell)) for cell in row] for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return rows or [["空表格"]]

    return [list(dataframe.columns)] + dataframe.astype(str).values.tolist()
